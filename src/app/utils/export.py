"""Shared Excel and PDF export utilities."""

import io
import logging
import tempfile
import os
from datetime import datetime, timedelta
from typing import List, Optional

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from fpdf import FPDF

logger = logging.getLogger(__name__)

# Brand colors
TEAL = (13, 148, 136)
TEAL_600 = (13, 148, 136)
TEAL_400 = (45, 212, 191)
TEAL_200 = (153, 246, 228)
TEAL_50 = (240, 253, 250)
TEAL_LIGHT = (240, 253, 250)
SLATE_900 = (15, 23, 42)
SLATE_700 = (51, 65, 85)
SLATE_500 = (100, 116, 139)
SLATE_400 = (148, 163, 184)
SLATE_300 = (203, 213, 225)
SLATE_100 = (241, 245, 249)
SLATE_50 = (248, 250, 252)
WHITE = (255, 255, 255)
EMERALD = (16, 185, 129)
RED = (239, 68, 68)
RED_400 = (248, 113, 113)
AMBER = (245, 158, 11)
AMBER_400 = (251, 191, 36)
BLUE = (59, 130, 246)
INDIGO = (99, 102, 241)
VIOLET = (139, 92, 246)
ORANGE = (249, 115, 22)

# Accent palette used by the UI KPI cards (name -> (icon bg/text, value bar))
ACCENTS = {
    "teal": TEAL, "red": RED, "violet": VIOLET, "amber": AMBER,
    "blue": BLUE, "indigo": INDIGO, "orange": ORANGE, "slate": SLATE_500,
    "emerald": EMERALD,
}

# Cache downloaded images during a single PDF generation
_image_cache: dict[str, str] = {}


def _download_image(url: str) -> Optional[str]:
    """Download image to a temp file, return path. Cached per URL."""
    if not url or url in ("", "-"):
        return None
    if url in _image_cache:
        return _image_cache[url]
        
    try:
        resp = requests.get(url, timeout=8, stream=True)
        if resp.status_code != 200:
            return None
        suffix = ".jpg" if "jpg" in url.lower() or "jpeg" in url.lower() else ".png"
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
        for chunk in resp.iter_content(4096):
            tmp.write(chunk)
        tmp.close()
        _image_cache[url] = tmp.name
        return tmp.name
    except Exception:
        return None


def _cleanup_image_cache():
    """Remove temp files after PDF generation."""
    for path in _image_cache.values():
        try:
            os.unlink(path)
        except OSError:
            pass
    _image_cache.clear()


def generate_excel(
    title: str,
    headers: List[str],
    rows: List[List],
    filename: str,
) -> io.BytesIO:
    """Generate an Excel file with styled headers and auto-width columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row in rows:
            if col_idx - 1 < len(row):
                max_len = max(max_len, len(str(row[col_idx - 1] or "")))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_excel_multi(
    sheets: List[tuple],
) -> io.BytesIO:
    """Generate a multi-sheet Excel workbook.

    `sheets` is a list of (title, headers, rows) tuples — one worksheet per
    entry. Styling mirrors `generate_excel` (teal header, borders, auto-width,
    frozen header row).
    """
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    wb = Workbook()
    # Remove the default sheet; we add our own named ones.
    default_ws = wb.active
    wb.remove(default_ws)

    used_titles = set()
    for raw_title, headers, rows in sheets:
        # Excel sheet titles: max 31 chars, must be unique, no special chars.
        title = (raw_title or "Sheet")[:31]
        base, n = title, 1
        while title in used_titles:
            suffix = f" {n}"
            title = base[: 31 - len(suffix)] + suffix
            n += 1
        used_titles.add(title)

        ws = wb.create_sheet(title=title)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row in rows:
                if col_idx - 1 < len(row):
                    max_len = max(max_len, len(str(row[col_idx - 1] if row[col_idx - 1] is not None else "")))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

        if headers:
            ws.freeze_panes = "A2"

    # Safety: a workbook must have at least one sheet.
    if not wb.sheetnames:
        wb.create_sheet(title="Report")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


class ParkingPDF(FPDF):
    """Branded PDF for AI Parking reports."""

    def __init__(self, title: str = "AI Parking Report", orientation: str = "P"):
        super().__init__(orientation=orientation)
        self.report_title = title

    def header(self):
        # Title only on the first page; continuation pages just carry the table
        # (a small top margin, no repeated title).
        if self.page_no() > 1:
            self.set_y(12)
            return
        # Clean header — dark title on white (no teal banner). The green status
        # pill drawn by the report body carries the "Updated …" timestamp.
        self.set_font("Helvetica", "B", 15)
        self.set_text_color(*SLATE_900)
        self.set_xy(10, 7)
        self.cell(0, 8, self.report_title, align="L")
        self.set_y(17)

    def footer(self):
        self.set_y(-12)
        self.set_font("Helvetica", "", 7)
        self.set_text_color(*SLATE_500)
        self.cell(95, 8, "AI Parking Management System", align="L")
        self.cell(95, 8, f"Page {self.page_no()}/{{nb}}", align="R")

    def _wrap_text(self, text, w, font_size, style="B"):
        """Split text into lines fitting width `w` (hard-splits over-long words)."""
        self.set_font("Helvetica", style, font_size)
        words = str(text).split()
        lines, cur = [], ""
        for word in words:
            trial = (cur + " " + word).strip()
            if self.get_string_width(trial) <= w - 1:
                cur = trial
                continue
            if cur:
                lines.append(cur)
                cur = ""
            if self.get_string_width(word) > w - 1:
                part = ""
                for ch in word:
                    if self.get_string_width(part + ch) > w - 1 and part:
                        lines.append(part)
                        part = ch
                    else:
                        part += ch
                cur = part
            else:
                cur = word
        if cur:
            lines.append(cur)
        return lines or ["-"]

    def _draw_card(self, fields: List[tuple], image_url: str = "", record_num: int = 0):
        """Record card: image on the left (~70%), data on the right (~30%).

        The image is aspect-fit inside its panel; the data column uses a stacked
        label-over-value layout with wrapping, so labels and values are never cut.
        """
        card_x = 10
        card_w = 190
        card_h = 124
        pad = 5
        gap = 6
        inner_w = card_w - 2 * pad - gap
        img_w = round(inner_w * 0.70)            # 70% image
        data_w = inner_w - img_w                  # 30% data
        img_h = card_h - 16

        if self.get_y() + card_h > 275:
            self.add_page()
        card_y = self.get_y()

        # Card background
        self.set_fill_color(*WHITE)
        self.set_draw_color(*SLATE_300)
        self.set_line_width(0.2)
        self.rect(card_x, card_y, card_w, card_h, "FD")

        # Record number badge (top-left)
        self.set_fill_color(*TEAL)
        self.set_font("Helvetica", "B", 7)
        self.set_text_color(*WHITE)
        self.set_xy(card_x + pad, card_y + 3)
        self.cell(9, 5, f"#{record_num}", align="C", fill=True)

        # ── Image panel (left ~70%) — aspect-fit, centered, NO border ──
        img_x = card_x + pad
        img_y = card_y + 10
        img_path = _download_image(image_url)
        drawn = False
        if img_path:
            try:
                self.image(img_path, x=img_x, y=img_y, w=img_w, h=img_h, keep_aspect_ratio=True)
                drawn = True
            except TypeError:  # older fpdf without keep_aspect_ratio
                try:
                    self.image(img_path, img_x, img_y, img_w, img_h)
                    drawn = True
                except Exception:
                    drawn = False
            except Exception:
                drawn = False
        if not drawn:
            self._draw_image_placeholder(img_x, img_y, img_w, img_h)

        # ── Data panel (right ~30%) — stacked label / value, wrapped (never cut) ──
        data_x = img_x + img_w + gap
        label_h = 4.8
        val_line_h = 5.0
        field_gap = 1.4

        prepared = []
        for label, value in fields:
            val_str = str(value) if value not in (None, "") else "-"
            lines = self._wrap_text(val_str, data_w, 9.5, "B")
            prepared.append((label, val_str, lines))

        total_h = sum(label_h + len(lines) * val_line_h + field_gap for _, _, lines in prepared) - field_gap
        cy = card_y + max(6, (card_h - total_h) / 2)

        for label, val_str, lines in prepared:
            color = SLATE_900
            if val_str in ("-", "N/A"):
                color = SLATE_400
            elif val_str in ("Active", "Still Parked", "IN"):
                color = TEAL
            elif val_str in ("Completed",):
                color = EMERALD
            elif val_str in ("OUT", "Yes"):
                color = RED

            # Label (uppercase, muted)
            self.set_xy(data_x, cy)
            self.set_font("Helvetica", "B", 8)
            self.set_text_color(*SLATE_500)
            self.cell(data_w, label_h, _fit_text(self, _latin1(str(label)).upper(), data_w), align="L")
            cy += label_h

            # Value (bold, wrapped)
            self.set_font("Helvetica", "B", 9.5)
            self.set_text_color(*color)
            for ln in lines:
                self.set_xy(data_x, cy)
                self.cell(data_w, val_line_h, ln, align="L")
                cy += val_line_h
            cy += field_gap

        self.set_y(card_y + card_h + 4)

    def _draw_image_placeholder(self, x, y, w, h):
        """Draw a gray placeholder when image is unavailable."""
        self.set_fill_color(*WHITE)
        self.set_draw_color(*SLATE_300)
        self.rect(x, y, w, h, "FD")
        cx = x + w / 2
        cy = y + h / 2
        self.set_draw_color(*SLATE_300)
        self.rect(cx - 8, cy - 6, 16, 12, "D")
        self.line(cx - 5, cy + 3, cx - 2, cy - 2)
        self.line(cx - 2, cy - 2, cx + 2, cy + 1)
        self.line(cx + 2, cy + 1, cx + 5, cy - 1)
        self.ellipse(cx + 3, cy - 4, 3, 3, "D")
        self.set_font("Helvetica", "I", 6)
        self.set_text_color(*SLATE_300)
        self.set_xy(x, y + h - 8)
        self.cell(w, 6, "No image", align="C")


def generate_pdf_table(
    title: str,
    headers: List[str],
    rows: List[List],
    col_widths: Optional[List[int]] = None,
) -> io.BytesIO:
    """Generate a PDF with a styled table."""
    pdf = ParkingPDF(title)
    pdf.alias_nb_pages()
    pdf.add_page()

    if not col_widths:
        available = 190
        col_widths = [available // len(headers)] * len(headers)

    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(*TEAL)
    pdf.set_text_color(*WHITE)
    for i, header in enumerate(headers):
        pdf.cell(col_widths[i], 7, header, border=1, fill=True, align="C")
    pdf.ln()

    pdf.set_font("Helvetica", "", 7)
    pdf.set_text_color(*SLATE_900)
    for row_idx, row in enumerate(rows):
        if row_idx % 2 == 0:
            pdf.set_fill_color(*SLATE_100)
        else:
            pdf.set_fill_color(*WHITE)
        for i, val in enumerate(row):
            w = col_widths[i] if i < len(col_widths) else 20
            pdf.cell(w, 6, str(val or "")[:50], border=1, fill=True, align="C")
        pdf.ln()

    output = io.BytesIO()
    pdf.output(output)
    output.seek(0)
    return output


_UNICODE_FALLBACKS = {
    0x2014: "-", 0x2013: "-",          # em / en dash
    0x2018: "'", 0x2019: "'",          # curly single quotes
    0x201C: '"', 0x201D: '"',          # curly double quotes
    0x2022: "*", 0x2026: "...",        # bullet, ellipsis
    0x00A0: " ",                        # non-breaking space
}


def _latin1(text: str) -> str:
    """FPDF core fonts are latin-1 only — transliterate common punctuation,
    then drop anything still unrepresentable."""
    return str(text).translate(_UNICODE_FALLBACKS).encode("latin-1", "replace").decode("latin-1")


def _fit_text(pdf, text: str, width: float) -> str:
    """Truncate text (with '..') so it fits within `width` mm."""
    text = _latin1(str(text))
    if pdf.get_string_width(text) <= width - 2:
        return text
    while text and pdf.get_string_width(text + "..") > width - 2:
        text = text[:-1]
    return (text + "..") if text else ""


def generate_pdf_multi(
    title: str,
    sections: List[tuple],
    max_rows_per_section: int = 200,
) -> io.BytesIO:
    """Multi-section landscape PDF — one titled table per section.

    `sections` is a list of (section_title, headers, rows) tuples, mirroring
    generate_excel_multi. Column widths auto-size to content; long cells are
    truncated; the header row repeats after every page break.
    """
    pdf = ParkingPDF(title, orientation="L")
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=14)
    pdf.add_page()
    usable = pdf.w - 20  # 10mm margins each side

    def draw_header_row(headers, widths):
        pdf.set_font("Helvetica", "B", 7)
        pdf.set_fill_color(*TEAL)
        pdf.set_text_color(*WHITE)
        for i, h in enumerate(headers):
            pdf.cell(widths[i], 6, _fit_text(pdf, h, widths[i]), border=1, fill=True, align="C")
        pdf.ln()

    for sec_title, headers, rows in sections:
        if pdf.get_y() > pdf.h - 40:
            pdf.add_page()
        pdf.ln(2)
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_text_color(*TEAL)
        pdf.cell(0, 7, _latin1(sec_title), new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(*SLATE_900)
        if not headers:
            continue

        n = len(headers)
        maxlens = [len(str(h)) for h in headers]
        for r in rows[:80]:
            for i in range(n):
                v = r[i] if i < len(r) else ""
                maxlens[i] = max(maxlens[i], len(str(v if v is not None else "")))
        capped = [min(m, 40) for m in maxlens]
        tot = sum(capped) or 1
        widths = [max(11, usable * c / tot) for c in capped]
        scale = usable / sum(widths)
        widths = [w * scale for w in widths]

        draw_header_row(headers, widths)

        pdf.set_font("Helvetica", "", 6)
        pdf.set_text_color(*SLATE_900)
        shown = rows[:max_rows_per_section]
        for ri, row in enumerate(shown):
            if pdf.get_y() > pdf.h - 16:
                pdf.add_page()
                draw_header_row(headers, widths)
                pdf.set_font("Helvetica", "", 6)
                pdf.set_text_color(*SLATE_900)
            pdf.set_fill_color(*(SLATE_100 if ri % 2 == 0 else WHITE))
            for i in range(n):
                v = row[i] if i < len(row) else ""
                pdf.cell(widths[i], 5, _fit_text(pdf, v if v is not None else "", widths[i]), border=1, fill=True, align="C")
            pdf.ln()

        if len(rows) > max_rows_per_section:
            pdf.set_font("Helvetica", "I", 6)
            pdf.set_text_color(*SLATE_500)
            pdf.cell(0, 5, f"... {len(rows) - max_rows_per_section} more rows (see CSV/Excel export)", new_x="LMARGIN", new_y="NEXT")

    output = io.BytesIO()
    pdf.output(output)
    output.seek(0)
    return output


# ─────────────────────────────────────────────────────────────────────────────
# Styled "dashboard-like" report PDF — mirrors the frontend Reports page design.
# ─────────────────────────────────────────────────────────────────────────────
def _txt(pdf, x, y, w, text, size=8, style="", color=SLATE_900, align="L", h=5):
    pdf.set_xy(x, y)
    pdf.set_font("Helvetica", style, size)
    pdf.set_text_color(*color)
    pdf.cell(w if w > 0 else 0, h, _fit_text(pdf, text, w) if w > 0 else _latin1(str(text)), align=align)


def _section_title(pdf, x, y, text):
    _txt(pdf, x, y, 0, text, size=11, style="B", color=SLATE_900, h=7)
    return y + 9


def _panel(pdf, x, y, w, h, title):
    pdf.set_draw_color(*SLATE_300)
    pdf.set_fill_color(*WHITE)
    pdf.rect(x, y, w, h, "FD")
    if title:
        _txt(pdf, x + 3, y + 2.5, w - 6, title, size=8, style="B", color=SLATE_700)


def _kpi_card(pdf, x, y, w, h, label, value, sub, accent):
    pdf.set_draw_color(*SLATE_300)
    pdf.set_fill_color(*WHITE)
    pdf.rect(x, y, w, h, "FD")
    pdf.set_fill_color(*accent)
    pdf.rect(x, y, w, 1.6, "F")              # top accent bar
    pdf.rect(x + 3, y + 3.5, 5, 5, "F")      # icon chip
    _txt(pdf, x + 2, y + 9, w - 4, str(value), size=13, style="B", color=SLATE_900, h=6)
    _txt(pdf, x + 2, y + 15.8, w - 4, label.upper(), size=6, style="B", color=SLATE_400, h=3.5)
    if sub:
        _txt(pdf, x + 2, y + 19, w - 4, sub, size=6, style="", color=SLATE_400, h=3)


def _kpi_row(pdf, x, y, w, cards, h=24, gap=3):
    n = len(cards) or 1
    cw = (w - (n - 1) * gap) / n
    for i, (lbl, val, sub, acc) in enumerate(cards):
        _kpi_card(pdf, x + i * (cw + gap), y, cw, h, lbl, val, sub, ACCENTS.get(acc, TEAL))
    return y + h


def _split_bar(pdf, x, y, w, h, segments):
    total = sum(v for v, _ in segments) or 1
    cx = x
    for v, color in segments:
        sw = w * v / total
        if sw > 0:
            pdf.set_fill_color(*color)
            pdf.rect(cx, y, sw, h, "F")
            cx += sw
    # legend
    lx = x
    for v, color in segments:
        pdf.set_fill_color(*color)
        pdf.rect(lx, y + h + 2, 2.5, 2.5, "F")
        _txt(pdf, lx + 3.5, y + h + 1.5, 28, str(v), size=7, color=SLATE_500)
        lx += 34


def _mini_metrics(pdf, x, y, w, items):
    n = len(items) or 1
    cw = w / n
    for i, (lbl, val, color) in enumerate(items):
        _txt(pdf, x + i * cw, y, cw, str(val), size=12, style="B", color=color, align="C", h=6)
        _txt(pdf, x + i * cw, y + 6, cw, lbl, size=6, color=SLATE_400, align="C", h=4)


def _vehicle_block(pdf, x, y, w, label, occupied, available, total, accent):
    pdf.set_draw_color(*SLATE_300)
    pdf.rect(x, y, w, 26, "D")
    _txt(pdf, x + 2, y + 1.5, w - 14, label, size=8, style="B", color=SLATE_700)
    _txt(pdf, x + 2, y + 1.5, w - 4, f"{total} total", size=6, color=SLATE_400, align="R")
    _txt(pdf, x + 2, y + 9, (w - 4) / 2, str(occupied), size=13, style="B", color=RED, align="C", h=7)
    _txt(pdf, x + 2, y + 18, (w - 4) / 2, "Occupied", size=6, color=SLATE_400, align="C", h=3)
    _txt(pdf, x + 2 + (w - 4) / 2, y + 9, (w - 4) / 2, str(available), size=13, style="B", color=EMERALD, align="C", h=7)
    _txt(pdf, x + 2 + (w - 4) / 2, y + 18, (w - 4) / 2, "Available", size=6, color=SLATE_400, align="C", h=3)


def _mini_stat(pdf, x, y, w, h, label, value):
    pdf.set_draw_color(*SLATE_300)
    pdf.set_fill_color(*WHITE)
    pdf.rect(x, y, w, h, "FD")
    _txt(pdf, x + 3, y + h / 2 - 2.5, w * 0.6, label, size=7, color=SLATE_500)
    _txt(pdf, x + w * 0.4 - 3, y + h / 2 - 3, w * 0.6, str(value), size=9, style="B", color=SLATE_900, align="R")


def _hourly_bars(pdf, x, y, w, h, data, color):
    mx = max(data) if data and max(data) > 0 else 1
    bw = w / max(len(data), 1)
    pdf.set_fill_color(*color)
    for i, v in enumerate(data):
        bh = (v / mx) * h
        if bh > 0.3:
            pdf.rect(x + i * bw, y + h - bh, bw * 0.72, bh, "F")
    pdf.set_draw_color(*SLATE_300)
    pdf.line(x, y + h, x + w, y + h)
    for hh in range(0, 24, 3):
        _txt(pdf, x + hh * bw - 2, y + h + 0.5, 8, str(hh), size=5, color=SLATE_400, align="L", h=3)


def _hbar_list(pdf, x, y, w, items, color, row_h=5.5):
    mx = items[0]["count"] if items else 1
    for i, it in enumerate(items):
        ry = y + i * row_h
        _txt(pdf, x, ry, 6, str(i + 1), size=6, style="B", color=SLATE_400, align="R", h=row_h - 1)
        _txt(pdf, x + 7, ry, 26, it["label"], size=7, style="B", color=SLATE_700, h=row_h - 1)
        bar_x = x + 35
        bar_w = w - 35 - 10
        pdf.set_fill_color(*SLATE_100)
        pdf.rect(bar_x, ry + 0.7, bar_w, row_h - 2, "F")
        pdf.set_fill_color(*color)
        pdf.rect(bar_x, ry + 0.7, bar_w * (it["count"] / (mx or 1)), row_h - 2, "F")
        _txt(pdf, x + w - 10, ry, 10, str(it["count"]), size=7, style="B", color=SLATE_600 if False else SLATE_500, align="R", h=row_h - 1)


def _occ_color(pct, threshold):
    if pct >= 95:
        return RED, WHITE
    if pct >= threshold:
        return TEAL_600, WHITE
    if pct >= threshold * 0.75:
        return TEAL_400, WHITE
    if pct >= 30:
        return TEAL_200, SLATE_700
    if pct > 0:
        return TEAL_50, SLATE_500
    return SLATE_50, SLATE_300


def _heatmap(pdf, x, y, w, zones, threshold):
    label_w = 52
    avg_w = 9
    grid_w = w - label_w - avg_w
    cell_w = grid_w / 24
    row_h = 5.2
    # header
    _txt(pdf, x, y, label_w, "Zone", size=6, style="B", color=SLATE_400)
    _txt(pdf, x + label_w, y, avg_w, "Avg", size=6, style="B", color=SLATE_400, align="C")
    for hh in range(0, 24, 2):
        _txt(pdf, x + label_w + avg_w + hh * cell_w, y, cell_w * 2, str(hh), size=5, color=SLATE_400, align="C", h=4)
    cy = y + 5
    for z in zones:
        _txt(pdf, x, cy + 0.5, label_w - 1, z["zone_name"], size=6, style="B", color=SLATE_700, h=4)
        _txt(pdf, x, cy + 4, label_w - 1, (z.get("location_name") or "")[:34], size=5, color=SLATE_400, h=3)
        af, at = _occ_color(z["avg_occupancy_pct"], threshold)
        pdf.set_fill_color(*af)
        pdf.rect(x + label_w, cy, avg_w - 1, row_h, "F")
        _txt(pdf, x + label_w, cy + 1, avg_w - 1, f"{round(z['avg_occupancy_pct'])}", size=5, style="B", color=at, align="C", h=3)
        for hb in z["hourly_breakdown"]:
            f, _ = _occ_color(hb["occupancy_pct"], threshold)
            pdf.set_fill_color(*f)
            pdf.rect(x + label_w + avg_w + hb["hour"] * cell_w, cy, cell_w - 0.3, row_h, "F")
        cy += row_h + 1.5
    return cy


def _table(pdf, x, headers, rows, widths, max_rows=120, top=26):
    def draw_header():
        pdf.set_x(x)
        pdf.set_font("Helvetica", "B", 6.5)
        pdf.set_fill_color(*TEAL)
        pdf.set_text_color(*WHITE)
        for i, hh in enumerate(headers):
            pdf.cell(widths[i], 5.5, _fit_text(pdf, str(hh), widths[i]), border=1, fill=True, align="C")
        pdf.ln()

    draw_header()
    pdf.set_font("Helvetica", "", 6)
    for ri, row in enumerate(rows[:max_rows]):
        if pdf.get_y() > pdf.h - 14:
            pdf.add_page()
            pdf.set_y(top)
            draw_header()
            pdf.set_font("Helvetica", "", 6)
        pdf.set_x(x)
        pdf.set_fill_color(*(SLATE_50 if ri % 2 else WHITE))
        pdf.set_text_color(*SLATE_900)
        for i in range(len(headers)):
            v = row[i] if i < len(row) else ""
            pdf.cell(widths[i], 5, _fit_text(pdf, v if v is not None else "", widths[i]), border=1, fill=True, align="C")
        pdf.ln()
    if len(rows) > max_rows:
        pdf.set_x(x)
        _txt(pdf, x, pdf.get_y(), 0, f"... {len(rows) - max_rows} more rows (see Excel export)", size=6, style="I", color=SLATE_500)


def generate_report_pdf(title: str, range_label: str, d: dict) -> io.BytesIO:
    """Render a dashboard-styled report PDF that mirrors the frontend Reports page."""
    pdf = ParkingPDF(title, orientation="L")
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(False)
    M = 10
    W = pdf.w - 2 * M

    # ── Page 1: Overview ──
    pdf.add_page()
    _txt(pdf, M, 20, W, f"Period: {range_label}", size=8, color=SLATE_500)
    y = _section_title(pdf, M, 25, "Overview")
    y = _kpi_row(pdf, M, y, W, d["kpis"]) + 5

    pw = (W - 6) / 2
    ph = 42
    _panel(pdf, M, y, pw, ph, "AI Parking - Live Slots")
    sc = d["slot_counts"]
    _split_bar(pdf, M + 4, y + 11, pw - 8, 6, [(sc["available"], EMERALD), (sc["occupied"], RED_400), (sc["obstructed"], AMBER_400)])
    _mini_metrics(pdf, M + 4, y + 26, pw - 8, [
        ("Total", sc["total"], SLATE_900), ("Available", sc["available"], EMERALD),
        ("Occupied", sc["occupied"], RED), ("Obstructed", sc["obstructed"], AMBER),
    ])
    ax = M + pw + 6
    _panel(pdf, ax, y, pw, ph, "ANPR - Live Occupancy")
    a = d["anpr_summary"]
    vbw = (pw - 12) / 2
    _vehicle_block(pdf, ax + 4, y + 11, vbw, "Car", a["car_occupied"], a["car_available"], a["car_total"], BLUE)
    _vehicle_block(pdf, ax + 8 + vbw, y + 11, vbw, "2-Wheeler", a["two_wheeler_occupied"], a["two_wheeler_available"], a["two_wheeler_total"], INDIGO)
    y += ph + 6

    hl = d["highlights"]
    n2 = len(hl) or 1
    gap = 3
    cw2 = (W - (n2 - 1) * gap) / n2
    for i, (lbl, val) in enumerate(hl):
        _mini_stat(pdf, M + i * (cw2 + gap), y, cw2, 12, lbl, val)

    # ── Page 2: AI Parking ──
    pdf.add_page()
    p = d["parking"]
    y = _section_title(pdf, M, 24, "AI Parking")
    y = _kpi_row(pdf, M, y, W, p["stats"], h=22) + 5

    lpw = W * 0.58
    rpw = W - lpw - 6
    ph = 46
    _panel(pdf, M, y, lpw, ph, "Hourly Activity")
    _hourly_bars(pdf, M + 4, y + 9, lpw - 8, ph - 16, p["hourly"], TEAL)
    _panel(pdf, M + lpw + 6, y, rpw, ph, "Duration Breakdown")
    dur = p["duration"]
    dur_items = [("< 30 min", dur.get("under_30m", 0), EMERALD), ("30m-1h", dur.get("30m_to_1h", 0), TEAL_400),
                 ("1-2 h", dur.get("1h_to_2h", 0), BLUE), ("2-8 h", dur.get("2h_to_8h", 0), AMBER_400),
                 ("> 8 h", dur.get("over_8h", 0), RED_400)]
    dmx = max([v for _, v, _ in dur_items] + [1])
    rx = M + lpw + 6 + 4                      # inside the right (Duration) panel
    bar_x = rx + 22
    bar_w = rpw - 8 - 22 - 12
    dy = y + 11
    for lbl, v, color in dur_items:
        _txt(pdf, rx, dy, 20, lbl, size=6, color=SLATE_500, align="R", h=5)
        pdf.set_fill_color(*SLATE_100)
        pdf.rect(bar_x, dy + 0.5, bar_w, 4.5, "F")
        pdf.set_fill_color(*color)
        pdf.rect(bar_x, dy + 0.5, bar_w * (v / dmx), 4.5, "F")
        _txt(pdf, bar_x + bar_w + 1, dy, 10, str(v), size=6, style="B", color=SLATE_500, align="L", h=5)
        dy += 6.5
    y += ph + 5

    _panel(pdf, M, y, lpw, 38, "Most Active Slots")
    _hbar_list(pdf, M + 4, y + 9, lpw - 8, p["top_slots"][:5], BLUE)
    _panel(pdf, M + lpw + 6, y, rpw, 38, "Devices & Alerts")
    dv = p["device"]
    _mini_metrics(pdf, M + lpw + 10, y + 12, rpw - 8, [
        ("Total", dv["total"], SLATE_900), ("Online", dv["online"], EMERALD), ("Offline", dv["offline"], RED),
    ])
    al = p["alert"]
    _txt(pdf, M + lpw + 10, y + 24, rpw - 8, f"Alerts: {al['total']} total - {al['active']} active - {al['resolved']} resolved", size=7, color=SLATE_500)
    y += 38 + 6

    _txt(pdf, M, y, W, "Parking Sessions", size=9, style="B", color=SLATE_800 if False else SLATE_900, h=6)
    pdf.set_y(y + 6)
    ps_w = [14, 18, 20, 26, 46, 34, 36, 36, 20, 22]  # sum 272 <= 277
    _table(pdf, M, ["Slot", "Type", "Vehicle", "Area", "Location", "Camera", "Entry", "Exit", "Dur(min)", "Status"], p["sessions"], ps_w, max_rows=80)

    # ── Page 3: ANPR ──
    pdf.add_page()
    an = d["anpr"]
    y = _section_title(pdf, M, 24, "ANPR")
    y = _kpi_row(pdf, M, y, W, an["kpis"], h=22) + 5

    lpw = W * 0.4
    rpw = W - lpw - 6
    ph = 44
    _panel(pdf, M, y, lpw, ph, "Vehicle Type Split")
    _split_bar(pdf, M + 4, y + 11, lpw - 8, 6, [(an["cars"], BLUE), (an["two_wheelers"], INDIGO)])
    tot_v = (an["cars"] + an["two_wheelers"]) or 1
    _txt(pdf, M + 4, y + 24, (lpw - 8) / 2, f"{round(an['cars'] / tot_v * 100)}%  {an['cars']} cars", size=8, style="B", color=BLUE, align="C")
    _txt(pdf, M + 4 + (lpw - 8) / 2, y + 24, (lpw - 8) / 2, f"{round(an['two_wheelers'] / tot_v * 100)}%  {an['two_wheelers']} 2W", size=8, style="B", color=INDIGO, align="C")
    _panel(pdf, M + lpw + 6, y, rpw, ph, "Hourly Entry Pattern")
    _hourly_bars(pdf, M + lpw + 10, y + 9, rpw - 8, ph - 16, an["hourly"], VIOLET)
    y += ph + 5

    half = (W - 6) / 2
    _panel(pdf, M, y, half, 38, "Top Frequent Plates")
    _hbar_list(pdf, M + 4, y + 9, half - 8, an["top_plates"][:5], VIOLET)
    _panel(pdf, M + half + 6, y, half, 38, "Busiest Locations")
    _hbar_list(pdf, M + half + 10, y + 9, half - 8, an["top_locations"][:5], TEAL)
    y += 38 + 6

    if an["locations"]:
        _txt(pdf, M, y, W, "Per-Location Occupancy", size=9, style="B", color=SLATE_900, h=6)
        pdf.set_y(y + 6)
        loc_w = [66, 32, 32, 32, 32, 30, 30]  # sum 254 <= 277
        _table(pdf, M, ["Location", "Car (Occ)", "Car (Tot)", "2W (Occ)", "2W (Tot)", "Occ %", "Avail %"], an["locations"], loc_w, max_rows=40)

    pdf.add_page()
    _txt(pdf, M, 24, W, "ANPR Sessions", size=9, style="B", color=SLATE_900, h=6)
    pdf.set_y(30)
    as_w = [40, 22, 64, 46, 46, 24, 18]  # sum 260 <= 277
    _table(pdf, M, ["Number Plate", "Type", "Location", "Entry", "Exit", "Duration", "Status"], an["sessions"], as_w, max_rows=90)

    # ── Page 4: Peak Occupancy ──
    pdf.add_page()
    occ = d["occupancy"]
    y = _section_title(pdf, M, 24, "Peak Occupancy")
    y = _kpi_row(pdf, M, y, W, occ["stats"], h=22) + 5
    if occ["zones"]:
        _panel(pdf, M, y, W, 12 + len(occ["zones"]) * 6.7, "Occupancy Heatmap")
        _heatmap(pdf, M + 4, y + 9, W - 8, occ["zones"], occ["threshold"])
        y += 12 + len(occ["zones"]) * 6.7 + 6
        if occ["insights"]:
            _txt(pdf, M, y, W, "Peak Occupancy Insights", size=9, style="B", color=SLATE_900, h=6)
            iy = y + 7
            for ins in occ["insights"][:8]:
                if iy > pdf.h - 14:
                    pdf.add_page()
                    iy = 26
                _txt(pdf, M + 2, iy, W - 4, "- " + ins, size=7, color=SLATE_700, h=5)
                iy += 5.5
    else:
        _txt(pdf, M, y + 4, W, "No occupancy data for the selected period.", size=9, color=SLATE_400)

    output = io.BytesIO()
    pdf.output(output)
    output.seek(0)
    return output


def generate_pdf_with_images(
    title: str,
    records: List[dict],
) -> io.BytesIO:
    """Generate a PDF with card-style records including real images."""
    try:
        pdf = ParkingPDF(title)
        pdf.alias_nb_pages()
        pdf.add_page()

        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(*SLATE_500)
        pdf.cell(0, 6, f"Total Records: {len(records)}", align="L", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        for idx, record in enumerate(records, 1):
            pdf._draw_card(
                fields=record.get("fields", []),
                image_url=record.get("image_url", ""),
                record_num=idx,
            )

        output = io.BytesIO()
        pdf.output(output)
        output.seek(0)
        return output
    finally:
        _cleanup_image_cache()


# ─────────────────────────────────────────────────────────────────────────────
# Live Parking Status PDF — mirrors the public/shared "view" page design:
# summary cards (Cars / Two-Wheeler) on top, then per-camera image cards.
# ─────────────────────────────────────────────────────────────────────────────
def _tint(color, f=0.86):
    """Lighten a color toward white by factor f (0..1)."""
    return tuple(int(ci + (255 - ci) * f) for ci in color)


def _draw_vehicle_icon(pdf, ix, cy, is_car, color, s=1.0):
    """Car / two-wheeler icon, vertically centered on cy. Base width ~7mm * s."""
    pdf.set_draw_color(*color)
    pdf.set_fill_color(*color)
    if is_car:
        pdf.set_line_width(0.3 * s)
        pdf.rect(ix, cy - 0.6 * s, 7 * s, 2.4 * s, "F", round_corners=True, corner_radius=0.7 * s)
        pdf.rect(ix + 1.7 * s, cy - 2.4 * s, 3.4 * s, 2.0 * s, "F", round_corners=True, corner_radius=0.6 * s)
        pdf.ellipse(ix + 0.8 * s, cy + 1.4 * s, 1.7 * s, 1.7 * s, "F")
        pdf.ellipse(ix + 4.5 * s, cy + 1.4 * s, 1.7 * s, 1.7 * s, "F")
    else:
        pdf.set_line_width(0.5 * s)
        pdf.ellipse(ix, cy + 0.2 * s, 2.8 * s, 2.8 * s, "D")
        pdf.ellipse(ix + 4.2 * s, cy + 0.2 * s, 2.8 * s, 2.8 * s, "D")
        pdf.line(ix + 1.4 * s, cy + 1.6 * s, ix + 4.0 * s, cy - 0.6 * s)
        pdf.line(ix + 4.0 * s, cy - 0.6 * s, ix + 5.6 * s, cy + 1.6 * s)
        pdf.line(ix + 4.0 * s, cy - 0.6 * s, ix + 5.2 * s, cy - 0.8 * s)


def _summary_card(pdf, x, y, w, h, title, accent, occupied, available, total):
    """Count card matching the UI: tinted card, centered title + icon, 3 divided columns."""
    pdf.set_fill_color(*_tint(accent, 0.93))
    pdf.set_draw_color(*_tint(accent, 0.55))
    pdf.set_line_width(0.2)
    try:
        pdf.rect(x, y, w, h, "FD", round_corners=True, corner_radius=2.5)
    except TypeError:
        pdf.rect(x, y, w, h, "FD")

    header_h = 9
    # Combined icon + title, left-aligned in the header (smaller, balanced)
    s = 0.85
    title_size = 10.5
    icon_w = 7 * s
    gap = 2.0
    gx = x + 5
    icy = y + header_h / 2
    _draw_vehicle_icon(pdf, gx, icy, title.strip().lower().startswith("car"), accent, s=s)
    _txt(pdf, gx + icon_w + gap, icy - 2.6, w - (icon_w + gap) - 7, title, size=title_size, style="B", color=accent, h=5)

    # Divider under header
    pdf.set_draw_color(*_tint(accent, 0.5))
    pdf.set_line_width(0.2)
    pdf.line(x + 3, y + header_h, x + w - 3, y + header_h)

    # Three columns (label on top, big number below), separated by vertical dividers
    sy = y + header_h
    sh = h - header_h
    col_w = w / 3
    for i, (lbl, val, c) in enumerate([
        ("OCCUPIED", occupied, RED), ("AVAILABLE", available, EMERALD), ("TOTAL", total, accent),
    ]):
        cx = x + i * col_w
        if i > 0:
            pdf.set_draw_color(*_tint(accent, 0.5))
            pdf.line(cx, sy + 3, cx, y + h - 3)
        # Vertically center the label+value pair within the stats area
        _txt(pdf, cx, sy + sh * 0.26, col_w, lbl, size=6.5, style="B", color=SLATE_500, align="C", h=4)
        _txt(pdf, cx, sy + sh * 0.56, col_w, str(val), size=14, style="B", color=c, align="C", h=7)


def _count_table(pdf, x, y, w, h, title, accent, occupied, available, total):
    """A 'Cars' / 'Two Wheeler' count card: vehicle icon + title + Status/Count rows."""
    pdf.set_fill_color(*WHITE)
    pdf.set_draw_color(*SLATE_300)
    pdf.set_line_width(0.2)
    pdf.rect(x, y, w, h, "FD")
    pdf.set_fill_color(*accent)
    pdf.rect(x, y, w, 1.4, "F")                       # top accent line

    # Vehicle icon + title centered and vertically aligned
    is_car = title.strip().lower().startswith("car")
    s = 0.7
    pdf.set_font("Helvetica", "B", 9)
    text_w = pdf.get_string_width(title)
    icon_w = 7 * s
    gap = 1.5
    group_w = icon_w + gap + text_w
    gx = x + (w - group_w) / 2
    icy = y + 7.2
    _draw_vehicle_icon(pdf, gx, icy, is_car, accent, s=s)
    _txt(pdf, gx + icon_w + gap, icy - 2.5, text_w + 4, title, size=9, style="B", color=accent, h=5)

    ry = y + 12
    row_h = (y + h - ry - 1) / 3
    for i, (lbl, val, c) in enumerate([
        ("Occupied", occupied, RED), ("Available", available, EMERALD), ("Total", total, accent),
    ]):
        cy = ry + i * row_h
        if i > 0:
            pdf.set_draw_color(*SLATE_100)
            pdf.set_line_width(0.2)
            pdf.line(x + 3, cy, x + w - 3, cy)
        _txt(pdf, x + 3, cy + row_h / 2 - 2.5, w * 0.55, lbl, size=8, color=SLATE_700, h=5)
        _txt(pdf, x, cy + row_h / 2 - 3, w - 3, str(val), size=14, style="B", color=c, align="R", h=6)


def generate_public_view_pdf(view: dict) -> io.BytesIO:
    """Live parking status report: summary cards + per-camera image cards."""
    try:
        pdf = ParkingPDF(view.get("name") or "Live Parking Status", orientation="P")
        pdf.alias_nb_pages()
        pdf.set_auto_page_break(False)
        M = 10
        W = pdf.w - 2 * M
        pdf.add_page()

        _txt(pdf, M, 20, W, "LIVE PARKING STATUS", size=8, style="B", color=SLATE_400)

        # ── Summary cards (Cars / Two-Wheeler) ──
        s = view["summary"]
        y = 26
        cw = (W - 6) / 2
        _summary_card(pdf, M, y, cw, 34, "Cars", BLUE,
                      s["car_occupied"], s["car_available"], s["car_total"])
        _summary_card(pdf, M + cw + 6, y, cw, 34, "Two Wheeler", INDIGO,
                      s["tw_occupied"], s["tw_available"], s["tw_total"])
        y += 34 + 7

        # ── Per-camera image cards ──
        cameras = view.get("cameras", [])
        if not cameras:
            _txt(pdf, M, y + 6, W, "No camera feeds available.", size=10, color=SLATE_400)

        for cam in cameras:
            content_h = 86
            card_h = 11.5 + content_h + 5
            if y + card_h > pdf.h - 14:
                pdf.add_page()
                y = 26

            pdf.set_fill_color(*WHITE)
            pdf.set_draw_color(*SLATE_300)
            pdf.set_line_width(0.2)
            pdf.rect(M, y, W, card_h, "FD")

            # Header strip (camera label + location)
            pdf.set_fill_color(*TEAL_50)
            pdf.rect(M, y, W, 9.5, "F")
            pdf.set_fill_color(*TEAL)
            pdf.rect(M, y, 1.6, 9.5, "F")
            _txt(pdf, M + 4, y + 2.4, W * 0.6, cam.get("label") or "Camera", size=10, style="B", color=SLATE_900)
            _txt(pdf, M + 4, y + 2.4, W - 8, cam.get("location_name") or "", size=8, color=SLATE_500, align="R")

            # Content row: image (left ~64%) + Cars / Two-Wheeler tables (right)
            cy = y + 12
            ix = M + 4
            iw = W * 0.62
            rx = ix + iw + 5
            rw = (M + W) - rx - 4

            img = _download_image(cam.get("image_url"))
            drawn = False
            if img:
                try:
                    pdf.image(img, x=ix, y=cy, w=iw, h=content_h, keep_aspect_ratio=True)
                    drawn = True
                except TypeError:
                    try:
                        pdf.image(img, ix, cy, iw, content_h)
                        drawn = True
                    except Exception:
                        drawn = False
                except Exception:
                    drawn = False
            if not drawn:
                pdf._draw_image_placeholder(ix, cy, iw, content_h)

            # Right: two stacked count tables
            t_gap = 4
            t_h = (content_h - t_gap) / 2
            car, tw = cam["car"], cam["tw"]
            _count_table(pdf, rx, cy, rw, t_h, "Cars", BLUE, car["o"], car["a"], car["t"])
            _count_table(pdf, rx, cy + t_h + t_gap, rw, t_h, "Two Wheeler", INDIGO, tw["o"], tw["a"], tw["t"])

            y += card_h + 6

        output = io.BytesIO()
        pdf.output(output)
        output.seek(0)
        return output
    finally:
        _cleanup_image_cache()


# LEGACY card-per-scan layout — kept for easy rollback; not currently wired.
def generate_parking_history_pdf_legacy(items: list, title: str = "AI Parking History Report") -> io.BytesIO:
    """Portrait layout matching public share view:
    header strip (location + datetime), image left ~78%, count tables right.

    Each item: {datetime, location, camera, image_url, car:{o,a,t}, tw:{o,a,t}}.
    """
    try:
        pdf = ParkingPDF(title, orientation="P")
        pdf.alias_nb_pages()
        pdf.set_auto_page_break(False)
        M = 10
        W = pdf.w - 2 * M
        pdf.add_page()
        _txt(pdf, M, 20, W, f"Total Records: {len(items)}", size=9, color=SLATE_500)
        y = 27

        for it in items:
            content_h = 110
            header_h = 9.5
            card_h = header_h + content_h + 5
            if y + card_h > pdf.h - 14:
                pdf.add_page()
                y = 22

            # Card border
            pdf.set_fill_color(*WHITE)
            pdf.set_draw_color(*SLATE_300)
            pdf.set_line_width(0.2)
            pdf.rect(M, y, W, card_h, "FD")

            # Header strip: location (left), date & time (right)
            pdf.set_fill_color(*TEAL_50)
            pdf.rect(M, y, W, header_h, "F")
            pdf.set_fill_color(*TEAL)
            pdf.rect(M, y, 1.6, header_h, "F")
            _txt(pdf, M + 5, y + 2.4, W * 0.6, it.get("location") or "-", size=10, style="B", color=SLATE_900)
            _txt(pdf, M + 4, y + 2.4, W - 8, it.get("datetime") or "-", size=8.5, style="B", color=SLATE_500, align="R")

            # Content: image left ~76%, count tables right (flush, no gap)
            cy = y + header_h + 2
            ix = M + 3
            iw = W * 0.76
            rx = ix + iw + 1
            rw = (M + W) - rx - 2

            # Image
            drawn = False
            img = _download_image(it.get("image_url"))
            if img:
                try:
                    pdf.image(img, x=ix, y=cy, w=iw, h=content_h, keep_aspect_ratio=True)
                    drawn = True
                except TypeError:
                    try:
                        pdf.image(img, ix, cy, iw, content_h)
                        drawn = True
                    except Exception:
                        drawn = False
                except Exception:
                    drawn = False
            if not drawn:
                pdf._draw_image_placeholder(ix, cy, iw, content_h)

            # Right: two stacked count tables
            t_gap = 4
            t_h = (content_h - t_gap) / 2
            car, tw = it["car"], it["tw"]
            _count_table(pdf, rx, cy, rw, t_h, "Cars", BLUE, car["o"], car["a"], car["t"])
            _count_table(pdf, rx, cy + t_h + t_gap, rw, t_h, "Two Wheeler", INDIGO, tw["o"], tw["a"], tw["t"])

            y += card_h + 6

        output = io.BytesIO()
        pdf.output(output)
        output.seek(0)
        return output
    finally:
        _cleanup_image_cache()


def _anpr_val_color(v: str):
    if v in ("Inside", "Active", "Still Parked"):
        return TEAL
    if v in ("Exited", "Completed"):
        return SLATE_700
    return SLATE_900


def _detail_table(pdf, x, y, w, h, title, accent, is_car, rows):
    """Right-side detail panel: vehicle icon + title, then label/value rows."""
    pdf.set_fill_color(*WHITE)
    pdf.set_draw_color(*SLATE_300)
    pdf.set_line_width(0.2)
    pdf.rect(x, y, w, h, "FD")
    pdf.set_fill_color(*accent)
    pdf.rect(x, y, w, 1.4, "F")

    # Icon + title (centered)
    s = 0.7
    pdf.set_font("Helvetica", "B", 9)
    text_w = pdf.get_string_width(title)
    icon_w = 7 * s
    gap = 1.5
    gx = x + (w - (icon_w + gap + text_w)) / 2
    icy = y + 7
    _draw_vehicle_icon(pdf, gx, icy, is_car, accent, s=s)
    _txt(pdf, gx + icon_w + gap, icy - 2.5, text_w + 4, title, size=9, style="B", color=accent, h=5)

    # Rows: label left, value right (wrapped)
    ry = y + 12
    row_h = (y + h - ry - 2) / max(len(rows), 1)
    lw = w * 0.32
    vw = w - lw - 7
    for i, (lbl, val) in enumerate(rows):
        cyr = ry + i * row_h
        if i > 0:
            pdf.set_draw_color(*SLATE_100)
            pdf.set_line_width(0.2)
            pdf.line(x + 3, cyr, x + w - 3, cyr)
        _txt(pdf, x + 4, cyr + row_h / 2 - 2.2, lw, str(lbl).upper(), size=6, style="B", color=SLATE_400, h=4)
        vlines = pdf._wrap_text(str(val) if val not in (None, "") else "-", vw, 8, "B")[:2]
        vstart = cyr + row_h / 2 - (len(vlines) * 4) / 2
        for j, ln in enumerate(vlines):
            _txt(pdf, x + lw + 4, vstart + j * 4, vw, ln, size=8, style="B",
                 color=_anpr_val_color(str(val)), align="L", h=4)


# LEGACY card-per-session layout — kept for easy rollback; not currently wired.
def generate_anpr_sessions_pdf_legacy(items: list, title: str = "ANPR Sessions Report") -> io.BytesIO:
    """Same card design as parking history: header strip (location + datetime),
    image on the left, a detail panel (vehicle icon + fields) on the right.

    Each item: {location, datetime, image_url, type, is_car, rows:[(label,value)]}.
    """
    try:
        pdf = ParkingPDF(title, orientation="P")
        pdf.alias_nb_pages()
        pdf.set_auto_page_break(False)
        M = 10
        W = pdf.w - 2 * M
        pdf.add_page()
        _txt(pdf, M, 20, W, f"Total Records: {len(items)}", size=9, color=SLATE_500)
        y = 27

        for it in items:
            content_h = 110
            header_h = 9.5
            card_h = header_h + content_h + 5
            if y + card_h > pdf.h - 14:
                pdf.add_page()
                y = 22

            pdf.set_fill_color(*WHITE)
            pdf.set_draw_color(*SLATE_300)
            pdf.set_line_width(0.2)
            pdf.rect(M, y, W, card_h, "FD")

            # Header strip: location (left) + datetime (right)
            pdf.set_fill_color(*TEAL_50)
            pdf.rect(M, y, W, header_h, "F")
            pdf.set_fill_color(*TEAL)
            pdf.rect(M, y, 1.6, header_h, "F")
            _txt(pdf, M + 5, y + 2.4, W * 0.6, it.get("location") or "-", size=10, style="B", color=SLATE_900)
            _txt(pdf, M + 4, y + 2.4, W - 8, it.get("datetime") or "-", size=8.5, style="B", color=SLATE_500, align="R")

            # Content: image left, detail panel right
            cy = y + header_h + 2
            ix = M + 3
            iw = W * 0.64
            rx = ix + iw + 2
            rw = (M + W) - rx - 2

            img = _download_image(it.get("image_url"))
            drawn = False
            if img:
                try:
                    pdf.image(img, x=ix, y=cy, w=iw, h=content_h, keep_aspect_ratio=True)
                    drawn = True
                except TypeError:
                    try:
                        pdf.image(img, ix, cy, iw, content_h)
                        drawn = True
                    except Exception:
                        drawn = False
                except Exception:
                    drawn = False
            if not drawn:
                pdf._draw_image_placeholder(ix, cy, iw, content_h)

            accent = BLUE if it.get("is_car") else INDIGO
            _detail_table(pdf, rx, cy, rw, content_h, it.get("type") or "Vehicle",
                          accent, bool(it.get("is_car")), it.get("rows") or [])

            y += card_h + 6

        output = io.BytesIO()
        pdf.output(output)
        output.seek(0)
        return output
    finally:
        _cleanup_image_cache()


# ─────────────────────────────────────────────────────────────────────────────
# Parking occupancy report (table layout) — matches the web "Parking occupancy
# report" page: Cars/Bikes summary tiles + an occupancy-records table with
# per-row snapshot thumbnails. (Card-per-scan layout preserved as *_legacy.)
# ─────────────────────────────────────────────────────────────────────────────
def _occ_tile(pdf, x, y, w, h, label, value, kind="neutral"):
    """Flat summary tile (reference UI): soft tinted card, small label on top,
    big value below. kind: neutral (grey) / occupied (red) / available (green)
    / in (blue) / out (amber)."""
    if kind == "occupied":
        bg, lcol, vcol = _tint(RED, 0.93), RED, RED
    elif kind == "available":
        bg, lcol, vcol = _tint(EMERALD, 0.93), EMERALD, EMERALD
    elif kind == "in":
        bg, lcol, vcol = _tint(BLUE, 0.93), BLUE, BLUE
    elif kind == "out":
        bg, lcol, vcol = _tint(AMBER, 0.93), AMBER, AMBER
    else:
        bg, lcol, vcol = SLATE_50, SLATE_500, SLATE_900

    # Flat tinted card with a subtle rounded border (no accent bar, no dot).
    pdf.set_fill_color(*bg)
    pdf.set_draw_color(*_tint(lcol, 0.55))
    pdf.set_line_width(0.2)
    try:
        pdf.rect(x, y, w, h, "FD", round_corners=True, corner_radius=3)
    except TypeError:
        pdf.rect(x, y, w, h, "FD")

    pad = 6
    # Label (top-left)
    _txt(pdf, x + pad, y + 6, w - 2 * pad, label, size=8.5, style="B", color=lcol, h=4)
    # Big value (below label)
    _txt(pdf, x + pad, y + h - 13, w - 2 * pad, str(value), size=20, style="B", color=vcol, h=11)


def _report_head(pdf, meta):
    """Location subtitle (grey dot) + green status pill on the title row,
    matching the reference 'Kankaria … · Ahmedabad' / 'Live · updated' header."""
    M = 10
    W = pdf.w - 2 * M

    loc = meta.get("location") or ""
    if loc:
        pdf.set_fill_color(*SLATE_400)
        pdf.ellipse(M + 0.4, 17.3, 1.8, 1.8, "F")
        _txt(pdf, M + 4, 16.2, W * 0.62, loc, size=9, color=SLATE_500, h=4)

    status = meta.get("status") or ""
    if status:
        pdf.set_font("Helvetica", "B", 7)
        ph = 6.5
        pw = pdf.get_string_width(status) + 11
        px, py = M + W - pw, 7.5
        pdf.set_fill_color(*_tint(EMERALD, 0.85))
        pdf.set_draw_color(*_tint(EMERALD, 0.5))
        pdf.set_line_width(0.2)
        try:
            pdf.rect(px, py, pw, ph, "FD", round_corners=True, corner_radius=3)
        except TypeError:
            pdf.rect(px, py, pw, ph, "FD")
        pdf.set_fill_color(*EMERALD)
        pdf.ellipse(px + 4, py + ph / 2 - 1, 2, 2, "F")
        _txt(pdf, px + 7, py + 1.7, pw - 8, status, size=7, style="B", color=EMERALD, align="L", h=3.5)


def _table_frame(pdf, M, W, y0, y1, radius=2):
    """Subtle rounded outer border around a table block (single page span)."""
    pdf.set_draw_color(*SLATE_300)
    pdf.set_line_width(0.3)
    try:
        pdf.rect(M, y0, W, y1 - y0, "D", round_corners=True, corner_radius=radius)
    except TypeError:
        pdf.rect(M, y0, W, y1 - y0, "D")


def _records_table(pdf, rows, top=14, show_location=True):
    """Occupancy-records table with a snapshot thumbnail column (reference UI).

    show_location: include the Location column (for the All-locations report).
    When a single location is selected it's redundant (shown in the subtitle),
    so it's dropped and the freed width goes to the other columns.
    """
    M = 10
    W = pdf.w - 2 * M
    if show_location:
        headers = ["Snapshot", "Location", "Date & Time", "Car Occ", "Car Avail", "Car Total", "2W Occ", "2W Avail", "2W Total"]
        widths = [20, 34, 26, 18, 18, 18, 18, 19, 19]   # sum = 190
        aligns = ["L", "L", "L", "R", "R", "R", "R", "R", "R"]
    else:
        headers = ["Snapshot", "Date & Time", "Car Occ", "Car Avail", "Car Total", "2W Occ", "2W Avail", "2W Total"]
        widths = [20, 30, 23, 23, 23, 23, 24, 24]   # sum = 190
        aligns = ["L", "L", "R", "R", "R", "R", "R", "R"]
    row_h = 15

    def draw_header():
        hy = pdf.get_y()
        pdf.set_fill_color(*SLATE_50)
        pdf.rect(M, hy, W, 8, "F")
        cx = M
        for i, hh in enumerate(headers):
            if aligns[i] == "L":
                _txt(pdf, cx + 3, hy + 2.4, widths[i] - 5, hh, size=7.5, style="B", color=SLATE_500, align="L", h=3.5)
            else:
                _txt(pdf, cx, hy + 2.4, widths[i] - 3, hh, size=7.5, style="B", color=SLATE_500, align="R", h=3.5)
            cx += widths[i]
        pdf.set_draw_color(*SLATE_300)
        pdf.set_line_width(0.3)
        pdf.line(M, hy + 8, M + W, hy + 8)
        pdf.set_y(hy + 8)

    page_top = pdf.get_y()
    draw_header()
    for ri, r in enumerate(rows):
        if pdf.get_y() + row_h > pdf.h - 14:
            _table_frame(pdf, M, W, page_top, pdf.get_y())
            pdf.add_page()
            pdf.set_y(top)
            page_top = pdf.get_y()
            draw_header()
        ry = pdf.get_y()
        # Zebra background + row separator
        pdf.set_fill_color(*(SLATE_50 if ri % 2 else WHITE))
        pdf.rect(M, ry, W, row_h, "F")
        pdf.set_draw_color(*SLATE_100)
        pdf.set_line_width(0.2)
        pdf.line(M, ry + row_h, M + W, ry + row_h)

        # Snapshot thumbnail
        tx, ty = M + 3, ry + 2
        tw, th = widths[0] - 6, row_h - 4
        img = _download_image(r.get("snapshot_url"))
        drawn = False
        if img:
            try:
                pdf.image(img, x=tx, y=ty, w=tw, h=th, keep_aspect_ratio=True)
                drawn = True
            except TypeError:
                try:
                    pdf.image(img, tx, ty, tw, th)
                    drawn = True
                except Exception:
                    drawn = False
            except Exception:
                drawn = False
        if not drawn:
            pdf._draw_image_placeholder(tx, ty, tw, th)

        cx = M + widths[0]
        wi = 1  # index into widths after the snapshot column

        # Location (only for the All-locations report) — wrapped up to 3 lines.
        if show_location:
            loc_w = widths[wi]
            loc_lines = pdf._wrap_text(r.get("location", "") or "-", loc_w - 5, 7.5, "")[:3]
            lh = 3.8
            ly = ry + (row_h - len(loc_lines) * lh) / 2
            for ln in loc_lines:
                _txt(pdf, cx + 3, ly, loc_w - 4, ln, size=7.5, style="", color=SLATE_700, align="L", h=lh)
                ly += lh
            cx += loc_w
            wi += 1

        # Date & Time — stacked on two lines so neither value is ever truncated.
        dt_w = widths[wi]
        _txt(pdf, cx + 3, ry + 3.4, dt_w - 4, r.get("date", ""), size=9, style="B", color=SLATE_700, align="L", h=4)
        _txt(pdf, cx + 3, ry + 8.4, dt_w - 4, r.get("time", ""), size=8, style="", color=SLATE_500, align="L", h=4)
        cx += dt_w
        wi += 1

        # Numeric columns (Occ=red, Avl=green, Total=dark), right-aligned.
        vmid = ry + row_h / 2 - 2
        nums = [
            (r.get("occ_car", 0), RED),
            (r.get("avl_car", 0), EMERALD),
            (r.get("tot_car", 0), SLATE_900),
            (r.get("occ_bike", 0), RED),
            (r.get("avl_bike", 0), EMERALD),
            (r.get("tot_bike", 0), SLATE_900),
        ]
        for i, (val, color) in enumerate(nums):
            wcol = widths[wi + i]
            _txt(pdf, cx, vmid, wcol - 4, str(val), size=9, style="B", color=color, align="R", h=4)
            cx += wcol
        pdf.set_y(ry + row_h)

    _table_frame(pdf, M, W, page_top, pdf.get_y())


def generate_parking_history_pdf(meta: dict, summary: dict, rows: list) -> io.BytesIO:
    """Parking occupancy report: Cars/Bikes summary tiles + occupancy-records table.

    meta: {title, location, status}
    summary: {car:{total,occupied,available}, bike:{total,occupied,available}}
    rows: [{snapshot_url, date, time, occ_car, avl_car, occ_bike, avl_bike}]
    """
    try:
        pdf = ParkingPDF(meta.get("title") or "Parking Occupancy Report", orientation="P")
        pdf.alias_nb_pages()
        pdf.set_auto_page_break(False)
        M = 10
        W = pdf.w - 2 * M
        pdf.add_page()

        _report_head(pdf, meta)

        car = summary.get("car", {})
        bike = summary.get("bike", {})
        tw = (W - 2 * 6) / 3
        th = 24

        y = 25
        _txt(pdf, M, y, W, "Cars", size=10, style="B", color=SLATE_900, h=5)
        y += 7
        _occ_tile(pdf, M, y, tw, th, "Total cars", car.get("total", 0), "neutral")
        _occ_tile(pdf, M + tw + 6, y, tw, th, "Occupied", car.get("occupied", 0), "occupied")
        _occ_tile(pdf, M + 2 * (tw + 6), y, tw, th, "Available", car.get("available", 0), "available")
        y += th + 6

        _txt(pdf, M, y, W, "Bikes", size=10, style="B", color=SLATE_900, h=5)
        y += 7
        _occ_tile(pdf, M, y, tw, th, "Total bikes", bike.get("total", 0), "neutral")
        _occ_tile(pdf, M + tw + 6, y, tw, th, "Occupied", bike.get("occupied", 0), "occupied")
        _occ_tile(pdf, M + 2 * (tw + 6), y, tw, th, "Available", bike.get("available", 0), "available")
        y += th + 6

        # Overall occupancy card
        total_all = car.get("total", 0) + bike.get("total", 0)
        occ_all = car.get("occupied", 0) + bike.get("occupied", 0)
        avail_all = max(0, total_all - occ_all)
        occ_pct = round((occ_all / total_all * 100)) if total_all > 0 else 0
        bar_w_frac = occ_all / total_all if total_all > 0 else 0

        card_h = 20
        pdf.set_fill_color(*WHITE)
        pdf.set_draw_color(*SLATE_300)
        pdf.set_line_width(0.3)
        pdf.rect(M, y, W, card_h, "FD")
        col_w = W / 4
        # Total
        _txt(pdf, M + 4, y + 3, col_w, "TOTAL CAPACITY", size=6, style="B", color=SLATE_400)
        _txt(pdf, M + 4, y + 9, col_w, str(total_all), size=16, style="B", color=SLATE_900, h=8)
        # Occupied
        pdf.set_fill_color(254, 242, 242)  # red-50
        pdf.rect(M + col_w, y, col_w, card_h, "F")
        _txt(pdf, M + col_w + 4, y + 3, col_w, "OCCUPIED", size=6, style="B", color=RED)
        _txt(pdf, M + col_w + 4, y + 9, col_w, str(occ_all), size=16, style="B", color=RED, h=8)
        # Available
        pdf.set_fill_color(240, 253, 244)  # emerald-50
        pdf.rect(M + 2 * col_w, y, col_w, card_h, "F")
        _txt(pdf, M + 2 * col_w + 4, y + 3, col_w, "AVAILABLE", size=6, style="B", color=EMERALD)
        _txt(pdf, M + 2 * col_w + 4, y + 9, col_w, str(avail_all), size=16, style="B", color=EMERALD, h=8)
        # Occupancy %
        pdf.set_fill_color(*TEAL_50)
        pdf.rect(M + 3 * col_w, y, col_w, card_h, "F")
        _txt(pdf, M + 3 * col_w + 4, y + 3, col_w - 8, "OCCUPANCY", size=6, style="B", color=TEAL)
        _txt(pdf, M + 3 * col_w + 4, y + 9, col_w - 8, f"{occ_pct}%", size=16, style="B", color=TEAL, h=8)
        # Progress bar
        bar_y = y + card_h
        pdf.set_fill_color(209, 250, 229)  # emerald-100
        pdf.rect(M, bar_y, W, 2, "F")
        bar_color = RED if occ_pct >= 90 else AMBER if occ_pct >= 60 else TEAL
        pdf.set_fill_color(*bar_color)
        pdf.rect(M, bar_y, W * bar_w_frac, 2, "F")
        y += card_h + 8

        _txt(pdf, M, y, W, f"Occupancy records ({len(rows)})", size=10, style="B", color=SLATE_900, h=5)
        pdf.set_y(y + 6)
        _records_table(pdf, rows, show_location=meta.get("show_location", True))

        output = io.BytesIO()
        pdf.output(output)
        output.seek(0)
        return output
    finally:
        _cleanup_image_cache()


def _thumb_table(pdf, columns, rows, top=14):
    """Generic records table with a snapshot-thumbnail first column.

    columns: [{header, width, key, align, color}]. The column whose key is
    'snapshot_url' renders a thumbnail. `color` may be an RGB tuple, the string
    'status' (Inside/Active/Still Parked -> teal, else slate), or omitted.
    """
    M = 10
    W = pdf.w - 2 * M
    row_h = 15

    def draw_header():
        hy = pdf.get_y()
        pdf.set_fill_color(*SLATE_50)
        pdf.rect(M, hy, W, 8, "F")
        cx = M
        for c in columns:
            wcol = c["width"]
            if c.get("align", "L") == "L":
                _txt(pdf, cx + 3, hy + 2.4, wcol - 5, c["header"], size=7.5, style="B", color=SLATE_500, align="L", h=3.5)
            elif c.get("align") == "R":
                _txt(pdf, cx, hy + 2.4, wcol - 3, c["header"], size=7.5, style="B", color=SLATE_500, align="R", h=3.5)
            else:
                _txt(pdf, cx, hy + 2.4, wcol, c["header"], size=7.5, style="B", color=SLATE_500, align="C", h=3.5)
            cx += wcol
        pdf.set_draw_color(*SLATE_300)
        pdf.set_line_width(0.3)
        pdf.line(M, hy + 8, M + W, hy + 8)
        pdf.set_y(hy + 8)

    page_top = pdf.get_y()
    draw_header()
    for ri, r in enumerate(rows):
        if pdf.get_y() + row_h > pdf.h - 14:
            _table_frame(pdf, M, W, page_top, pdf.get_y())
            pdf.add_page()
            pdf.set_y(top)
            page_top = pdf.get_y()
            draw_header()
        ry = pdf.get_y()
        pdf.set_fill_color(*(SLATE_50 if ri % 2 else WHITE))
        pdf.rect(M, ry, W, row_h, "F")
        pdf.set_draw_color(*SLATE_100)
        pdf.set_line_width(0.2)
        pdf.line(M, ry + row_h, M + W, ry + row_h)

        cx = M
        vmid = ry + row_h / 2 - 2
        for c in columns:
            wcol = c["width"]
            key = c["key"]
            if key == "snapshot_url":
                tx, ty = cx + 3, ry + 2
                tw, th = wcol - 6, row_h - 4
                img = _download_image(r.get(key))
                drawn = False
                if img:
                    try:
                        pdf.image(img, x=tx, y=ty, w=tw, h=th, keep_aspect_ratio=True)
                        drawn = True
                    except TypeError:
                        try:
                            pdf.image(img, tx, ty, tw, th)
                            drawn = True
                        except Exception:
                            drawn = False
                    except Exception:
                        drawn = False
                if not drawn:
                    pdf._draw_image_placeholder(tx, ty, tw, th)
            elif c.get("stack"):
                # Two values stacked on two lines (e.g. Date over Time).
                k1, k2 = c["stack"]
                _txt(pdf, cx + 3, ry + 3.4, wcol - 4, str(r.get(k1, "")), size=8.5, style="B", color=SLATE_700, align="L", h=4)
                _txt(pdf, cx + 3, ry + 8.4, wcol - 4, str(r.get(k2, "")), size=8, style="", color=SLATE_500, align="L", h=4)
            elif c.get("wrap"):
                # Wrapped text (e.g. long Location names) — up to 3 lines.
                lines = pdf._wrap_text(str(r.get(key, "") or "-"), wcol - 5, 7.5, "")[:3]
                lh = 3.8
                ly = ry + (row_h - len(lines) * lh) / 2
                for ln in lines:
                    _txt(pdf, cx + 3, ly, wcol - 4, ln, size=7.5, style="", color=SLATE_700, align="L", h=lh)
                    ly += lh
            else:
                val = r.get(key, "")
                color = c.get("color")
                if color == "status":
                    col = TEAL if str(val) in ("Parked", "Inside", "Active", "Still Parked") else SLATE_700
                elif isinstance(color, tuple):
                    col = color
                else:
                    col = SLATE_700
                align = c.get("align", "L")
                if align == "L":
                    _txt(pdf, cx + 3, vmid, wcol - 4, str(val), size=8.5, style="B", color=col, align="L", h=4)
                elif align == "R":
                    _txt(pdf, cx, vmid, wcol - 4, str(val), size=8.5, style="B", color=col, align="R", h=4)
                else:
                    _txt(pdf, cx, vmid, wcol, str(val), size=8.5, style="B", color=col, align="C", h=4)
            cx += wcol
        pdf.set_y(ry + row_h)

    _table_frame(pdf, M, W, page_top, pdf.get_y())


def generate_anpr_sessions_pdf(meta: dict, summary: dict, rows: list) -> io.BytesIO:
    """ANPR sessions report (same design as the parking occupancy report):
    today's live parking-occupancy tiles (Total / Occupied / Available per
    vehicle type) + the ANPR sessions records table.

    meta: {title, location, status}
    summary: {car:{total,in,out,available}, bike:{total,in,out,available}}
    rows: [{snapshot_url, plate, type, date, in, out, duration, status}]
    """
    try:
        pdf = ParkingPDF(meta.get("title") or "ANPR Sessions Report", orientation="P")
        pdf.alias_nb_pages()
        pdf.set_auto_page_break(False)
        M = 10
        W = pdf.w - 2 * M
        pdf.add_page()

        _report_head(pdf, meta)

        car = summary.get("car", {})
        bike = summary.get("bike", {})
        tw = (W - 3 * 6) / 4   # four tiles per row: Total / In / Out / Available
        th = 24

        def tile_row(vals, total_label):
            _occ_tile(pdf, M, y, tw, th, total_label, vals.get("total", 0), "neutral")
            _occ_tile(pdf, M + (tw + 6), y, tw, th, "In", vals.get("in", 0), "in")
            _occ_tile(pdf, M + 2 * (tw + 6), y, tw, th, "Out", vals.get("out", 0), "out")
            _occ_tile(pdf, M + 3 * (tw + 6), y, tw, th, "Available", vals.get("available", 0), "available")

        y = 25
        _txt(pdf, M, y, W, "Cars", size=10, style="B", color=SLATE_900, h=5)
        y += 7
        tile_row(car, "Total cars")
        y += th + 6

        _txt(pdf, M, y, W, "Two Wheeler", size=10, style="B", color=SLATE_900, h=5)
        y += 7
        tile_row(bike, "Total bikes")
        y += th + 7

        _txt(pdf, M, y, W, f"Session records ({len(rows)})", size=10, style="B", color=SLATE_900, h=5)
        pdf.set_y(y + 6)
        if meta.get("show_location", False):
            # All-locations report: add a Location column. Date & Time and Out
            # are each stacked onto two lines (sum of widths = 190).
            columns = [
                {"header": "Snapshot", "width": 20, "key": "snapshot_url", "align": "L"},
                {"header": "Location", "width": 27, "key": "location", "align": "L", "wrap": True},
                {"header": "Plate", "width": 25, "key": "plate", "align": "L", "color": SLATE_900},
                {"header": "Type", "width": 26, "key": "type", "align": "L"},
                {"header": "In", "width": 24, "key": "date", "align": "L", "stack": ("date", "in")},
                {"header": "Out", "width": 24, "key": "out_date", "align": "L", "stack": ("out_date", "out_time")},
                {"header": "Duration", "width": 22, "key": "duration", "align": "C"},
                {"header": "Status", "width": 22, "key": "status", "align": "C", "color": "status"},
            ]
        else:
            # Single-location report: no Location column (it's in the subtitle).
            # Date & Time and Out are stacked onto two lines so nothing is cut.
            columns = [
                {"header": "Snapshot", "width": 26, "key": "snapshot_url", "align": "L"},
                {"header": "Plate", "width": 28, "key": "plate", "align": "L", "color": SLATE_900},
                {"header": "Type", "width": 30, "key": "type", "align": "L"},
                {"header": "In", "width": 28, "key": "date", "align": "L", "stack": ("date", "in")},
                {"header": "Out", "width": 28, "key": "out_date", "align": "L", "stack": ("out_date", "out_time")},
                {"header": "Duration", "width": 26, "key": "duration", "align": "C"},
                {"header": "Status", "width": 24, "key": "status", "align": "C", "color": "status"},
            ]
        _thumb_table(pdf, columns, rows)

        output = io.BytesIO()
        pdf.output(output)
        output.seek(0)
        return output
    finally:
        _cleanup_image_cache()
